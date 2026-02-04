from openpyxl.worksheet.worksheet import Worksheet

from util import (
    A_COL,
    B_COL,
    C_COL,
    E_COL,
    F_COL,
    J_COL,
    K_COL,
    convert_date,
    extract_number,
)


def get_individual_election_office(building: Worksheet):
    """家屋費（選挙事務所費）の個別データを取得する。

    Excelシートの4行目以降から、日付、金額、種別、目的、備考を取得する。
    日付セルがNoneになったら処理を終了する。

    Args:
        building (Worksheet): 家屋費のExcelシート。

    Returns:
        list[dict]: 選挙事務所費の個別データリスト。各要素は以下のキーを持つ辞書:
            - category (str): データの種類を表す名前（常に"building"）。
            - date (str): 日付（YYYY-MM-DD形式）。date_cell.valueがdatetime型であることを前提とする。
            - price (int or float): 金額。
            - type (str or None): 種別。
            - purpose (str or None): 目的。
            - non_monetary_basis (str or None): 金銭以外の見積もりの根拠。
            - note (str or None): 備考。
    """

    building_data = []

    min_row = 4
    for row in building.iter_rows(min_row=min_row, max_col=K_COL + 1):
        date_cell = row[A_COL]
        price_cell = row[C_COL]
        type_cell = row[E_COL]
        purpose_cell = row[F_COL]
        non_monetary_basis_cell = row[J_COL]
        note_cell = row[K_COL]
        # Noneになったら終了
        if date_cell.value is None:
            break

        building_data.append(
            {
                "category": "building",  # シート名をカテゴリとして使用
                "date": convert_date(date_cell.value),
                "price": extract_number(price_cell.value),  # 金額
                "type": type_cell.value,  # 種別
                "purpose": purpose_cell.value,  # 支出の目的
                # 金銭以外の見積もりの根拠
                "non_monetary_basis": non_monetary_basis_cell.value,
                "note": note_cell.value,  # 備考
            }
        )

    return building_data


def get_total_election_office(building: Worksheet):
    """家屋費（選挙事務所費）の合計データを取得する。

    Excelシートの16行目以降から、「立候補準備のための支出」「選挙運動のための支出」「計」の
    3行を動的に検索して取得する。位置は個別データの数によって変わるため、動的に取得する。

    Args:
        building (Worksheet): 家屋費のExcelシート。

    Returns:
        tuple[list[dict], int or float]: 合計データのリストとチェックサム（「計」の金額）のタプル。
            合計データの各要素は以下のキーを持つ辞書:
            - name (str): 項目名（「立候補準備のための支出」「選挙運動のための支出」「計」のいずれか）。
            - price (int or float): 金額。
    """

    total_building_data = []
    count = 0
    json_checksum = 0  # jsonファイルの検証に使用

    # 合計に関する記述は16行目より下にある
    min_row = 16

    for row in building.iter_rows(min_row=min_row, max_col=C_COL + 1):
        # 型をチェック
        value = row[B_COL].value
        if not isinstance(value, str):
            continue

        # B列が立候補準備のための支出、選挙運動のための支出、計のいずれでもなければスキップ
        value_str = value.strip().replace("　", "")
        if value_str not in ["立候補準備のための支出", "選挙運動のための支出", "計"]:
            continue

        # 3行取得したら終了
        if count == 3:
            break

        price_value = row[C_COL].value
        price_value = extract_number(price_value)
        total_building_data.append({"name": value_str, "price": price_value})
        count += 1
        if value_str == "計":
            json_checksum = price_value

    return total_building_data, json_checksum


def get_individual_meeting_venue(building: Worksheet):
    """集会会場費の個別データを取得する。

    Excelシートの20行目以降から「月　　日」という文字列を検索し、
    その2行下から個別データの取得を開始する。
    日付セルがdatetime型でない場合はスキップし、Noneになったら処理を終了する。

    Args:
        building (Worksheet): 家屋費のExcelシート。

    Returns:
        list[dict]: 集会会場費の個別データリスト。各要素は以下のキーを持つ辞書:
            - category (str): データの種類を表す名前（常に"building"）。
            - date (str): 日付（YYYY-MM-DD形式）。date_cell.valueがdatetime型であることを前提とする。
            - price (int or float): 金額。
            - type (str or None): 種別。
            - purpose (str or None): 目的。
            - non_monetary_basis (str or None): 金銭以外の見積もりの根拠。
            - note (str or None): 備考。
            「月　　日」が見つからない場合は空のリストを返す。
    """
    meeting_venue_data = []

    # 22行目以降からスタートする スタート位置を特定
    min_row = 20
    start_row = 0

    for i, row in enumerate(
        building.iter_rows(min_row=min_row, max_col=A_COL + 1), start=min_row
    ):
        date_cell = row[A_COL]
        if date_cell.value is not None and date_cell.value == "月　　日":
            start_row = i + 2
            break
    else:
        exit("家屋シートの集合会場費の開始位置が見つかりません。")

    # 特定した行以降からスタートする
    for row in building.iter_rows(min_row=start_row, max_col=K_COL + 1):
        date_cell = row[A_COL]  # 日付
        price_cell = row[C_COL]  # 金額
        type_cell = row[E_COL]  # 種別
        purpose_cell = row[F_COL]  # 支出の目的
        non_monetary_basis_cell = row[J_COL]  # 金銭以外の見積もりの根拠
        note_cell = row[K_COL]  # 備考

        # Noneになったら終了
        if date_cell.value is None:
            break

        meeting_venue_data.append(
            {
                "category": "building",  # シート名をカテゴリとして使用
                "date": convert_date(date_cell.value),
                "price": extract_number(price_cell.value),  # 金額
                "type": type_cell.value,  # 種別
                "purpose": purpose_cell.value,  # 支出の目的
                # 金銭以外の見積もりの根拠
                "non_monetary_basis": non_monetary_basis_cell.value,
                "note": note_cell.value,  # 備考
            }
        )

    return meeting_venue_data


def get_total_meeting_venue(building: Worksheet):
    """集会会場費の合計データを取得する。

    Excelシートの34行目以降から、「立候補準備のための支出」「選挙運動のための支出」「計」の
    3行を動的に検索して取得する。位置は個別データの数によって変わるため、動的に取得する。

    Args:
        building (Worksheet): 家屋費のExcelシート。

    Returns:
        tuple[list[dict], int or float]: 合計データのリストとチェックサム（「計」の金額）のタプル。
            合計データの各要素は以下のキーを持つ辞書:
            - name (str): 項目名（「立候補準備のための支出」「選挙運動のための支出」「計」のいずれか）。
            - price (int or float): 金額。
    """

    total_meeting_venue_data = []
    count = 0
    json_checksum = 0  # jsonファイルの検証に使用

    # 合計に関する記述は34行目より下にある
    min_row = 34

    for row in building.iter_rows(min_row=min_row, max_col=C_COL + 1):
        # 型をチェック
        value = row[B_COL].value
        if not isinstance(value, str):
            continue

        # B列が立候補準備のための支出、選挙運動のための支出、計のいずれでもなければスキップ
        value_str = value.strip().replace("　", "")
        if value_str not in ["立候補準備のための支出", "選挙運動のための支出", "計"]:
            continue

        # 3行取得したら終了
        if count == 3:
            break

        price_value = row[C_COL].value
        price_value = extract_number(price_value)
        total_meeting_venue_data.append({"name": value_str, "price": price_value})
        count += 1
        if value_str == "計":
            json_checksum = price_value

    return total_meeting_venue_data, json_checksum


def get_building(building: Worksheet):
    """家屋費の全データを取得する。

    選挙事務所費と集会会場費の個別データおよび合計データを取得し、
    チェックサムを合算して1つの辞書にまとめて返す。

    Args:
        building (Worksheet): 家屋費のExcelシート。

    Returns:
        dict: 以下のキーを持つ辞書:
            - individual_election_office (list[dict]): 選挙事務所費の個別データリスト。
            - total_election_office (list[dict]): 選挙事務所費の合計データリスト。
            - individual_meeting_venue (list[dict]): 集会会場費の個別データリスト。
            - total_meeting_venue (list[dict]): 集会会場費の合計データリスト。
            - json_checksum (int or float): チェックサム（選挙事務所費と集会会場費の合計の合計）。
    """
    individual_election_office = get_individual_election_office(building)

    total_election_office, json_checksum1 = get_total_election_office(building)

    individual_meeting_venue = get_individual_meeting_venue(building)

    total_meeting_venue, json_checksum2 = get_total_meeting_venue(building)

    json_checksum = json_checksum1 + json_checksum2

    return {
        "individual_election_office": individual_election_office,
        "total_election_office": total_election_office,
        "individual_meeting_venue": individual_meeting_venue,
        "total_meeting_venue": total_meeting_venue,
        "json_checksum": json_checksum,
    }
