import re

from openpyxl.worksheet.worksheet import Worksheet

from util import A_COL, B_COL, C_COL, E_COL, I_COL, J_COL, convert_date, extract_number


def get_individual_income(income: Worksheet):
    """収入の部の個別データを取得する。

    Excelシートの4行目以降から、日付、金額、種別、備考を取得する。
    日付セルがNoneになったら処理を終了する。
    結合されているセルは、左端のセルに情報が入っている。

    Args:
        income (Worksheet): 収入の部のExcelシート。

    Returns:
        list[dict]: 収入の部のデータリスト。各要素は以下のキーを持つ辞書:
            - category (str): データの種類を表す名前（常に"income"）。
            - date (str): 日付（YYYY-MM-DD形式）。date_cell.valueがdatetime型であることを前提とする。
            - price (int or float): 金額。
            - type (str or None): 種別。
            - non_monetary_basis (str or None): 金銭以外の見積もりの根拠。
            - note (str or None): 備考。
    """
    income_data = []

    # 4行目以降, AからJの列を取得
    min_row = 4

    for row in income.iter_rows(min_row=min_row, max_col=J_COL + 1):
        date_cell = row[A_COL]  # 日付
        price_cell = row[C_COL]  # 金額
        type_cell = row[E_COL]  # 種別
        non_monetary_basis_cell = row[I_COL]  # 金銭以外の見積もりの根拠
        note_cell = row[J_COL]  # 備考
        # Noneになったら終了
        if date_cell.value is None:
            break

        income_data.append(
            {
                "category": "income",  # シート名をカテゴリとして使用
                "date": convert_date(date_cell.value),
                "price": extract_number(price_cell.value),  # 金額
                "type": type_cell.value,  # 種別
                # 金銭以外の見積もりの根拠
                "non_monetary_basis": non_monetary_basis_cell.value,
                "note": note_cell.value,  # 備考
            }
        )

    return income_data


def get_total_income(income: Worksheet):
    """総収入を取得する。

    Excelシートの7行目以降から、「寄附」「その他の収入」「計」「総計」の項目を
    動的に検索して取得する。合計に関する記述は9行あり、位置は個別データの数によって変わるため、
    動的に取得する。

    Args:
        income (Worksheet): 収入の部のExcelシート。

    Returns:
        list[dict]: 総収入のデータリスト。各要素は以下のキーを持つ辞書:
            - name (str): 項目名（「寄附」「その他の収入」「計」「総計」のいずれか）。
            - price (int or float): 金額。
    """

    total_income_data = []
    count = 0

    # 合計に関する記述は7行目より下にある
    min_row = 7

    for row in income.iter_rows(min_row=min_row, max_col=C_COL + 1):
        # B列が寄附, その他の収入, 計, 総計のいずれでもなければスキップ
        if row[B_COL].value not in ["寄附", "その他の収入", "計", "総計"]:
            continue
        # 9行取得したら終了
        if count == 9:
            break

        name_value = row[B_COL].value
        price_value = row[C_COL].value

        price_value = extract_number(price_value)
        total_income_data.append({"name": name_value, "price": price_value})
        count += 1

    return total_income_data


def get_public_expense_summary(income: Worksheet):
    """公費負担相当額のサマリーを取得する。

    Excelシートの18行目以降から「参考」という文字列を検索し、
    その行のB列から公費負担相当額の文字列を取得する。
    位置は個別データの数によって変わるため、動的に取得する。
    ここだけ文章で記述されているため、正規表現でパースして辞書形式で返す。

    Args:
        income (Worksheet): 収入の部のExcelシート。

    Returns:
        dict: 公費負担相当額のデータ。以下のキーを持つ:
            - summary (int): 公費負担相当額の総額。
            - breakdown (dict): 内訳の辞書。項目名をキー、金額を値とする。
            「参考」が見つからない場合は空の辞書を返す。
    """

    # 公費負担相当額に関する記述は7 + 9 + 2 = 18行目より下にある
    min_row = 18
    public_expense_summary_str = ""

    # ここではBしか取得しないため、row[0]で取得できる
    for row in income.iter_rows(min_row=min_row, max_col=B_COL + 1):
        if row[A_COL].value == "参考":
            public_expense_summary_str = row[B_COL].value
            break
    else:
        return {}

    # 総額を取得する正規表現
    summary_pattern = r"公費負担相当額\s+(\d+(?:,\d+)*)円"
    summary_match = re.search(summary_pattern, str(public_expense_summary_str))

    # 内訳を取得する正規表現
    breakdown_pattern = r"内訳\s+(.+)"
    breakdown_match = re.search(breakdown_pattern, str(public_expense_summary_str))

    public_expense_summary_data = {}

    # 総額を追加
    if summary_match:
        summary_amount = int(summary_match.group(1).replace(",", ""))
        public_expense_summary_data["summary"] = summary_amount

    # 内訳をパース
    if breakdown_match:
        breakdown_str = breakdown_match.group(1)
        # カンマを事前に削除
        breakdown_str = breakdown_str.replace(",", "")

        # 内訳内の各項目をパースする正規表現（項目名 + 数字 + 円 の形式）
        # スペースが無くても対応するため、数字の直前までを項目名とする
        item_pattern = r"([^0-9]+)(\d+)円"
        items = re.findall(item_pattern, breakdown_str)

        breakdown_data = {}
        for item_name, amount_str in items:
            # 項目名の前後のスペース、カンマ、読点を除去
            item_name = item_name.strip().replace("、", "").replace(",", "")
            # 金額を整数に変換
            amount = int(amount_str)
            breakdown_data[item_name] = amount

        public_expense_summary_data["breakdown"] = breakdown_data
    return public_expense_summary_data


def get_income(income: Worksheet):
    """収入の部の全データを取得する。

    個別の収入データ、総収入データ、公費負担相当額のサマリーを取得し、1つの辞書にまとめて返す。
    収入に関しては、前回計・総額などのdiffデータを取り扱っているため、checksumは用意していない。

    Args:
        income (Worksheet): 収入の部のExcelシート。

    Returns:
        dict: 以下のキーを持つ辞書:
            - individual_income (list[dict]): 収入の個別データリスト。
            - total_income (list[dict]): 総収入のデータリスト。
            - public_expense_summary (dict): 公費負担相当額のサマリーデータ。
    """
    individual_income = get_individual_income(income)

    total_income = get_total_income(income)

    public_expense_summary = get_public_expense_summary(income)

    return {
        "individual_income": individual_income,
        "total_income": total_income,
        "public_expense_summary": public_expense_summary,
    }
    # 収入に関しては、前回計・総額などのdiffデータを取り扱っているので、checksumは用意していない
